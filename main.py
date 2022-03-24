from pdf2image import convert_from_path
import cv2
import pytesseract
import tkinter as tk
from tkinter import HORIZONTAL, VERTICAL, messagebox
from functools import cmp_to_key
from openpyxl import load_workbook
from openpyxl.styles import Alignment
pytesseract.pytesseract.tesseract_cmd = r"Tesseract\tesseract.exe"

root = tk.Tk()

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

CurrRow = 1
CurrCol = 1
def write_to_excel(PrevY, CurrY, text, firstCnt):
    wb = load_workbook('Invoice.xlsx')
    sheet = wb.active #Reading active sheet.

    global CurrRow 
    global CurrCol

    if firstCnt == True:
        sheet.cell(CurrRow, CurrCol).value = text
        sheet.cell(CurrRow, CurrCol).alignment = Alignment(vertical='top')

    else:
        if isInRow(PrevY, CurrY):
            CurrCol += 1
            sheet.cell(CurrRow, CurrCol).value = text
            sheet.cell(CurrRow, CurrCol).alignment = Alignment(vertical='top')

        else:
            CurrRow += 1
            CurrCol = 1
            sheet.cell(CurrRow, CurrCol).value = text
            sheet.cell(CurrRow, CurrCol).alignment = Alignment(vertical='top')

    wb.save('Invoice.xlsx')

def isInRow(PreVal, CurVal):
    TOLERANCE = 20

    if(PreVal - TOLERANCE) <= CurVal <= ( PreVal + TOLERANCE ) : 
        return True
    else:
        return False

#Sorting function was obtained via https://stackoverflow.com/questions/65897395/how-can-i-sort-contours-from-left-to-right-and-top-to-bottom-in-a-consistent-man
def sorting_contours(contour_a, contour_b):

    contour_a = cv2.boundingRect(contour_a)
    contour_b = cv2.boundingRect(contour_b)

    if abs(contour_a[1] - contour_b[1]) <= 15:
        return contour_a[0] - contour_b[0]

    return contour_a[1] - contour_b[1]

def ROI(image_path):
    image = image_path
    #Converting images to grayscale
    if len(image.shape) != 2:
        image_to_gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        image_to_gray = image

    #Outlines the text
    blur = cv2.GaussianBlur(image_to_gray,(5,5),0)
    reference, thres = cv2.threshold(blur, 0, 255, cv2.THRESH_OTSU)
    thres = 255 - thres

    #Dilating so that the text can form one big blob. That way, the region selected is more accurate
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (20, 3))
    dilate = cv2.dilate(thres, kernel, iterations = 2)

    #Find contours, region's of interest,
    contours = cv2.findContours(dilate, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[0]

    sorted_contours = sorted(contours, key=cmp_to_key(sorting_contours))

    #text_to_copy = []
    firstCnt = False
    for i, cnt in enumerate(sorted_contours):
        x, y, w, h = cv2.boundingRect(cnt)

        if h > 20:
            cv2.rectangle(image, (x,y-3), (x+w, (y+h)+3), color = (255, 0, 255), thickness= 2)
            cropped = image_to_gray[y-3: (y + h) + 3, x: x + w]
            
            text = pytesseract.image_to_string(cropped, lang = 'eng', config='--psm 7 --oem 3')
            #text_to_copy.append(pytesseract.image_to_string(cropped, lang = 'eng', config='--psm 7 --oem 3'))

            if firstCnt == False:
                PrevY = y
                CurrY = y
                write_to_excel(PrevY, CurrY, text, firstCnt=True)
                firstCnt = True
            else:
                PrevY = CurrY
                CurrY = y
                write_to_excel(PrevY, CurrY, text, firstCnt=False)

    #joined_text = "".join(text_to_copy)
    #saveText(joined_text)

    return image

def saveText(text_to_copy):
    with open('text.txt', 'w') as f:
        f.writelines(text_to_copy)

def crop(naming_convention):
    im = cv2.imread(naming_convention)

    #Selecting region to Crop
    scale_percent = screen_height/screen_width * 0.42 # percent of original size
    width = int(im.shape[1] * scale_percent)
    height = int(im.shape[0] * scale_percent)
    dim = (width, height)

    imCopy = cv2.resize(im, dim)

    r = cv2.selectROI("Press C to cancel, Press Enter to process", imCopy, fromCenter=False, showCrosshair=False)

    #Cropping Image
    imCrop = im[int(r[1]/scale_percent):int((r[1]+r[3])/scale_percent), int(r[0]/scale_percent):int((r[0]+r[2])/scale_percent)]

    return imCrop

#Go through pages and save each one as an image and apply logic.
def main():
    #Storing all the pages of the PDF in a variable.
    images = convert_from_path(str(entry1.get()), 350, poppler_path=r"poppler-21.11.0\Library\bin")

    counter = 1
    for img in images:
        naming_convention = r"Images\Page " + str(counter) + ".jpg"
        img.save(naming_convention, 'JPEG')

        ROIimage = crop(naming_convention)
        ROIimage = ROI(ROIimage)
        
        cv2.imwrite(r"Images\Bounded " + str(counter) + ".jpg", ROIimage)

        counter+=1
    
    root.destroy()

tk.Label(root, text = "Enter File Path: ").grid(row = 0, sticky = tk.W)

entry1 = tk.Entry(root)
entry1.grid(row = 0, column = 1)

button1 = tk.Button(root, text = "Convert", command=main)
button1.grid(row = 0, column= 2, columnspan = 2, rowspan=2, padx = 5, pady = 5)

root.mainloop()